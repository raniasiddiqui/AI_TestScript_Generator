import streamlit as st
import re
import os
from datetime import datetime
import pandas as pd
from groq import Groq
import asyncio
import nest_asyncio
import sys
import json
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
nest_asyncio.apply()

# Page config
st.set_page_config(
    page_title="Playwright Test Runner & Reporter",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ────────────────────────────────────────────────
#               Custom CSS (unchanged)
# ────────────────────────────────────────────────
st.markdown("""
    <style>
    /* Main container */
    .main {
        background-color: #f8f9fa;
    }
    
    /* Headers */
    .main-header {
        font-size: 2rem;
        font-weight: 600;
        color: #2c3e50;
        text-align: center;
        margin-bottom: 1.5rem;
        padding: 1rem;
    }
    
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #f0f2f5;
        padding: 2rem 1rem;
    }
    
    [data-testid="stSidebar"] h2 {
        color: #2c3e50;
        font-size: 1.2rem;
        font-weight: 600;
        margin-bottom: 1rem;
    }
    
    [data-testid="stSidebar"] h3 {
        color: #34495e;
        font-size: 1rem;
        font-weight: 600;
        margin-top: 1.5rem;
        margin-bottom: 0.5rem;
    }
    
    [data-testid="stSidebar"] .element-container {
        margin-bottom: 0.5rem;
    }
    
    /* Info boxes */
    .info-box {
        background-color: #d1ecf1;
        border-left: 4px solid #0c5460;
        border-radius: 0.5rem;
        padding: 1.2rem;
        margin: 1.5rem 0;
    }
    
    .success-box {
        background-color: #d4edda;
        border-left: 4px solid #155724;
        border-radius: 0.5rem;
        padding: 1.2rem;
        margin: 1rem 0;
    }
    
    .section-container {
        background-color: white;
        border: 1px solid #e0e0e0;
        border-radius: 0.75rem;
        padding: 2rem;
        margin: 1.5rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .section-header {
        font-size: 1.5rem;
        font-weight: 600;
        color: #2c3e50;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    /* Form elements */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea {
        background-color: #ffffff !important;
        color: #2c3e50 !important;
        border: 1.5px solid #ced4da !important;
        border-radius: 0.5rem !important;
        padding: 0.75rem !important;
        font-size: 0.95rem !important;
    }
    
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #4a90e2 !important;
        box-shadow: 0 0 0 0.2rem rgba(74, 144, 226, 0.25) !important;
        outline: none !important;
    }
    
    .stTextInput label,
    .stTextArea label {
        color: #34495e !important;
        font-weight: 500 !important;
        font-size: 0.95rem !important;
        margin-bottom: 0.5rem !important;
    }
    
    ::placeholder {
        color: #95a5a6 !important;
        opacity: 1 !important;
    }
    
    /* Primary button */
    div[data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(135deg, #4a90e2 0%, #357abd 100%) !important;
        border: none !important;
        color: white !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        border-radius: 0.5rem !important;
        padding: 0.75rem 2rem !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 6px rgba(74, 144, 226, 0.3) !important;
    }
    
    div[data-testid="stFormSubmitButton"] > button:hover {
        background: linear-gradient(135deg, #357abd 0%, #2868a8 100%) !important;
        box-shadow: 0 6px 8px rgba(74, 144, 226, 0.4) !important;
        transform: translateY(-2px);
    }
    
    div[data-testid="stFormSubmitButton"] > button:active {
        transform: translateY(0);
    }
    
    /* Download button */
    .stDownloadButton > button {
        background-color: #27ae60 !important;
        border: none !important;
        color: white !important;
        font-weight: 600 !important;
        border-radius: 0.5rem !important;
        padding: 0.75rem 2rem !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
    }
    
    .stDownloadButton > button:hover {
        background-color: #229954 !important;
        box-shadow: 0 4px 6px rgba(39, 174, 96, 0.3) !important;
    }
    
    /* Code block styling */
    .stCodeBlock {
        background-color: #f8f9fa !important;
        border: 1px solid #e0e0e0 !important;
        border-radius: 0.5rem !important;
        margin: 1rem 0 !important;
    }
    
    /* Help text */
    .help-text {
        font-size: 0.85rem;
        color: #7f8c8d;
        margin-top: 0.25rem;
    }
    
    /* Divider */
    hr {
        margin: 2rem 0;
        border: none;
        border-top: 1px solid #e0e0e0;
    }
    
    /* Progress bar */
    .stProgress > div > div > div > div {
        background-color: #4a90e2;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background-color: #f8f9fa;
        border-radius: 0.5rem;
        font-weight: 500;
    }
    </style>
""", unsafe_allow_html=True)

# Groq setup
GROQ_API_KEY = st.secrets["groq_api_key"]
DEFAULT_GROQ_MODEL = st.secrets["groq_default_model"]

if not GROQ_API_KEY:
    st.error("❌ GROQ API key not found in config.json")
    st.stop()

groq_client = Groq(api_key=GROQ_API_KEY)

class GroqAgent:
    def __init__(self, system_prompt, model_name=DEFAULT_GROQ_MODEL):
        self.system_prompt = system_prompt
        self.model_name = model_name

    async def generate(self, user_content: str) -> str:
        try:
            completion = groq_client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {"role": "system", "content": self.system_prompt},
                    {"role": "user", "content": user_content}
                ],
                temperature=0.0,
                max_tokens=8000
            )
            return completion.choices[0].message.content.strip()
        except Exception as e:
            return f"Error: {str(e)}"

# ────────────────────────────────────────────────
#          ORIGINAL SCRIPT GENERATION PROMPT (unchanged)
# ────────────────────────────────────────────────
# TRANSFORM_PROMPT = """
# You are a strict Playwright instrumentation engine.

# You will receive ONE Playwright function.

# Your job is NOT to rewrite it.
# Your job is NOT to improve it.
# Your job is NOT to create a sample.
# Your job is ONLY to instrument it.

# STRICT RULES:

# 1. Copy the original function definition EXACTLY as provided.
# 2. Copy EVERY original line EXACTLY as-is.
# 3. DO NOT change indentation.
# 4. DO NOT change locators.
# 5. DO NOT change URLs.
# 6. DO NOT remove expect statements.
# 7. DO NOT modify function signature.
# 8. DO NOT create a new example.
# 9. DO NOT invent steps.

# You are ONLY allowed to:

# - Add try/except around EACH original executable statement
# - Add screenshot capture
# - Add step_logs.append(...)
# - Add step counter increment
# - Add a run() wrapper outside the function

# INSTRUMENTATION PATTERN:

# For each original statement:

# try:
#     ORIGINAL LINE HERE
#     page.screenshot(path=f"step_{{step_number}}_PASS.png")
#     step_logs.append(f"Step {{step_number}}: PASS")
# except Exception as e:
#     page.screenshot(path=f"step_{{step_number}}_FAIL.png")
#     step_logs.append(f"Step {{step_number}}: FAIL - {{str(e)}}")

# step_number += 1

# DO NOT merge steps.
# Each original line = one try block.

# After copying and instrumenting the function:

# Add:

# - step_logs = []
# - step_number = 1
# - run() function that:
#     - launches browser
#     - creates page
#     - calls original function
#     - writes Excel file

# Output ONLY Python code.

# Now instrument this function EXACTLY:

# {input_code}
# """

TRANSFORM_PROMPT = """
You are a strict Playwright instrumentation engine.

You will receive ONE Playwright function.

Your job is NOT to rewrite it.
Your job is NOT to improve it.
Your job is NOT to create a sample.
Your job is ONLY to instrument it and add validation
based on page.get_by_text() that already exists in the input.

STRICT RULES:

1. Copy the original function EXACTLY as provided.
2. Copy EVERY original line EXACTLY as-is.
3. DO NOT change indentation.
4. DO NOT change locators.
5. DO NOT change URLs.
6. DO NOT remove existing expect statements.
7. DO NOT modify function signature.
8. DO NOT invent new steps.
9. DO NOT change execution order.

You are ONLY allowed to:

- Add try/except around EACH original executable statement
- Add screenshot capture
- Add step_logs.append(...)
- Add step counter increment
- Add validation using page.get_by_text text from input
- Add run() wrapper

------------------------------------------------------------
SMART PAGE-BY-TEXT VALIDATION RULE (PARTIAL KEY MATCHING)
------------------------------------------------------------

1. First, scan the ORIGINAL function.
2. Extract any text used inside:
       expect(page.locator("body")).to_contain_text('')

3. Treat that text as EXPECTED_PAGE_TEXT.

4. Convert EXPECTED_PAGE_TEXT into partial keywords by:
   - Splitting text into words
   - Removing extra whitespace
   - Ignoring empty values
   - Converting to lowercase

5. After EACH successful original statement,
   add PARTIAL KEY MATCH validation using:

       page_text = page.locator("body").inner_text().lower()

       for keyword in EXPECTED_PAGE_TEXT.split():
           if keyword.strip():
               assert keyword.lower() in page_text, f"{{keyword}} not found in page"

6. This must:
   - Be inside try block
   - Come AFTER the original line
   - Not replace original expect statements
   - Not modify business logic

7. If multiple to_contain_text values exist,
   use the most recent one encountered above the current step.

8. If no to_contain_text exists in input,
   then use:
       expect(page.locator("body")).not_to_be_empty()

------------------------------------------------------------
INSTRUMENTATION PATTERN
------------------------------------------------------------

For each original statement:

try:
    ORIGINAL LINE HERE

    page_text = page.locator("body").inner_text().lower()
    for keyword in "ExtractedExpectedText".split():
        if keyword.strip():
            assert keyword.lower() in page_text, f"{{keyword}} not found in page"

    page.screenshot(path=f"step_{{step_number}}_PASS.png")
    step_logs.append(f"Step {{step_number}}: PASS")

except Exception as e:
    page.screenshot(path=f"step_{{step_number}}_FAIL.png")
    step_logs.append(f"Step {{step_number}}: FAIL - {{str(e)}}")

step_number += 1

------------------------------------------------------------

Each original line = one try block.
Do NOT merge steps.
Do NOT change logic.

After instrumenting:

Add:
- step_logs = []
- step_number = 1
- run() wrapper
- browser headless=False
- maximized
- Excel report writing

Output ONLY valid Python code.
No explanations.
No markdown.

Now instrument this function EXACTLY:

{input_code}
"""


# ────────────────────────────────────────────────
#     NEW: TEST CASE PLANNING PROMPT (inspired by PlannerOSS)
# ────────────────────────────────────────────────
TESTCASE_PLAN_PROMPT = """\
You are a test case documentation expert. Generate test cases for the provided Playwright script.

For context, use the flow mentioned in the original function, and generate test cases that cover the key features, user interactions, and expected outcomes in the script ONLY.

CRITICAL FORMATTING RULES:
1. Use EXACTLY this format for each test case (no variations)
2. Each field MUST start with "* " (asterisk + space)
3. Each field MUST be on a single line (no multi-line values except Step-by-step actions)
4. Test Case ID must be TC-1, TC-2, TC-3, etc. (sequential)
5. Separate test cases with ONE blank line

REQUIRED FORMAT PER TEST CASE:

* High Level Feature: [Category]
* Test Case ID: TC-[Number]
* Feature Name: [Specific feature of the script that is being tested. E.g. "Booking Flow", "Login Functionality", etc.]
* Test Scenario: [Summary of the user interaction or feature being tested in one line]
* Test Case: [Title]
* Test Case Description: [Details in one line of the feature being tested.]
* Step-by-step actions: [Single paragraph with all steps, no numbering]
* Possible Values: [Data or None]
* Sources: [Sources or None]
* Expected Result: [Pass criteria in one line]
* Data Correctness Checked: [Yes/No]
* Release/Platform Version: Web
* Automation Possibility: [Yes/No]
* Testing Type: [Type]
* Priority: [High/Medium/Low]

EXAMPLE:

 OUTPUT the test cases in the following format:
        STRICTLY ADHERE TO THIS FORMAT:
        - Test Case ID: TC-<number>
        - High Level Feature
        - Feature Name
        - Test Scenario
        - Test Case
        - Test Case Description
        - Step-by-step actions
        - Possible Values (if applicable, Type 'None' if there is none for a specific case)
        - Sources (if applicable, Type 'N/A' if there is none for a specific case)
        - Expected Result
        - Data Correctness Checked (if applicable, Type 'N/A' if there is none for a specific case)
        - Release/Platform Version (Web/Mobile/IOS/Android etc. If not applicable, write 'N/A')
        - Automation Possibility
        - Testing Type
        - Priority

NOW GENERATE TEST CASES FOR THIS SCRIPT:
```python
{{input_code}}
```

Extract flows from the provided script and generate test cases ONLY for flows that actually exist in the script.
Do NOT invent features.
Use actual button names, actual URLs, and actual assertions present in the script.
Generate all possible testcases that exist (20-25) ONLY FOR FLOWS THAT EXIST IN MY ORIGINAL SCRIPT THAT I INPUT. 
Donot repeat testcases, give me distinct ones for flows/functionality that exist in my original script.

"""

def clean_generated_code(raw: str) -> str:
    """
    Aggressive cleaning: remove ALL non-code content including test case docs.
    """
    lines = raw.splitlines()
    cleaned = []
    
    # Skip until we find the first import
    found_code_start = False
    
    for line in lines:
        stripped = line.strip()
        
        # Skip markdown code fences
        if stripped.startswith('```'):
            continue
            
        # Look for code start
        if not found_code_start:
            if stripped.startswith(('from ', 'import ')):
                found_code_start = True
            else:
                continue
        
        # After code starts, filter out test case documentation
        # Skip lines that are clearly test case docs (even as comments)
        if any(keyword in stripped.lower() for keyword in [
            'test case id:', 'tc-', 'high level feature:', 'feature name:',
            'test scenario:', 'test case:', 'test case description:',
            'step-by-step actions:', 'possible values:', 'sources:',
            'expected result:', 'data correctness checked:', 
            'release/platform version:', 'automation possibility:',
            'testing type:', 'priority:', '**test cases:**'
        ]):
            continue
            
        # Skip markdown-style bullets/headers (even in comments)
        if re.match(r'^\s*[#*\-•]\s*[A-Z]', stripped):
            continue
            
        # Skip numbered list items that look like test case headers
        if re.match(r'^\s*\d+\.\s*\*\*', stripped):
            continue
        
        # Keep the line if it passed all filters
        cleaned.append(line)
    
    result = '\n'.join(cleaned).strip()
    
    # Remove any remaining markdown artifacts
    result = re.sub(r'\*\*(.+?)\*\*', r'\1', result)  # Remove bold
    
    return result

def parse_and_export_testcases(test_cases_str: str):
    """
    Parse test cases with improved regex that handles the strict format.
    """
    # Split by test case blocks (look for "Test Case ID: TC-")
    test_blocks = re.split(r'\n(?=\* High Level Feature:)', test_cases_str.strip())
    
    all_data = []
    st.session_state.test_cases_list = []
    
    for block in test_blocks:
        if not block.strip():
            continue
            
        data = {
            'High Level Feature': '',
            'Test Case ID': '',
            'Feature Name': '',
            'Test Scenario': '',
            'Test Case': '',
            'Test Case Description': '',
            'Step-by-step actions': '',
            'Possible Values': '',
            'Sources': '',
            'Expected Result': '',
            'Data Correctness Checked': '',
            'Release/Platform Version': '',
            'Automation Possibility': '',
            'Testing Type': '',
            'Priority': ''
        }
        
        # Extract each field with simple, specific regex
        patterns = {
            'High Level Feature': r'\* High Level Feature:\s*(.+?)(?=\n\*|\Z)',
            'Test Case ID': r'\* Test Case ID:\s*(.+?)(?=\n\*|\Z)',
            'Feature Name': r'\* Feature Name:\s*(.+?)(?=\n\*|\Z)',
            'Test Scenario': r'\* Test Scenario:\s*(.+?)(?=\n\*|\Z)',
            'Test Case': r'\* Test Case:\s*(.+?)(?=\n\*|\Z)',
            'Test Case Description': r'\* Test Case Description:\s*(.+?)(?=\n\*|\Z)',
            'Step-by-step actions': r'\* Step-by-step actions:\s*(.+?)(?=\n\*|\Z)',
            'Possible Values': r'\* Possible Values:\s*(.+?)(?=\n\*|\Z)',
            'Sources': r'\* Sources:\s*(.+?)(?=\n\*|\Z)',
            'Expected Result': r'\* Expected Result:\s*(.+?)(?=\n\*|\Z)',
            'Data Correctness Checked': r'\* Data Correctness Checked:\s*(.+?)(?=\n\*|\Z)',
            'Release/Platform Version': r'\* Release/Platform Version:\s*(.+?)(?=\n\*|\Z)',
            'Automation Possibility': r'\* Automation Possibility:\s*(.+?)(?=\n\*|\Z)',
            'Testing Type': r'\* Testing Type:\s*(.+?)(?=\n\*|\Z)',
            'Priority': r'\* Priority:\s*(.+?)(?=\n\*|\Z)',
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, block, re.DOTALL | re.IGNORECASE)
            if match:
                value = match.group(1).strip()
                # Clean up any remaining formatting
                value = re.sub(r'\s+', ' ', value)  # Normalize whitespace
                data[key] = value
        
        # Only add if we found a Test Case ID
        if data['Test Case ID']:
            all_data.append(data)
            st.session_state.test_cases_list.append(data)
    
    # Export to Excel
    output_path = "cleaned_generated_test_cases.xlsx"
    if all_data:
        df = pd.DataFrame(all_data)
        
        try:
            df.to_excel(output_path, index=False)
            
            # Format Excel
            wb = load_workbook(output_path)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min((max_length + 2) * 1.2, 70)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
            st.success(f"✅ {len(all_data)} test cases exported to {output_path}")
            return True
        except Exception as e:
            st.error(f"Error saving Excel: {e}")
            return False
    else:
        st.warning("⚠️ No test cases parsed. Check LLM output format.")
        st.expander("Raw LLM Output").code(test_cases_str)
        return False
import streamlit as st
import asyncio
import pandas as pd
from datetime import datetime

# Assume these already exist in your project
# TRANSFORM_PROMPT
# TESTCASE_PLAN_PROMPT
# GroqAgent
# clean_generated_code
# parse_and_export_testcases


def main():
    st.markdown(
        '<div class="main-title">🤖 Automation Test Script Generation</div>',
        unsafe_allow_html=True
    )

    st.markdown(
    """
    <style>
    /* Light blue Generate button */
    div.stButton > button,
    div.stFormSubmitButton > button {
        background: #aee1ff !important;
        color: #00334d !important;
        font-weight: 600 !important;
        border-radius: 10px !important;
        height: 48px !important;
        border: none !important;
        box-shadow: none !important;
    }

    /* Hover state */
    div.stButton > button:hover,
    div.stFormSubmitButton > button:hover {
        background: #9fd8fb !important;
        color: #00334d !important;
    }

    /* Remove Streamlit primary gradient */
    div.stButton > button:focus,
    div.stFormSubmitButton > button:focus {
        background: #aee1ff !important;
        box-shadow: 0 0 0 2px rgba(174, 225, 255, 0.6) !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

    
    st.markdown(
        """
        <style>
        /* Page background */
        .stApp {
            background-color: #f6f7f9;
        }

        /* Main title */
        .main-title {
            text-align: center;
            font-size: 26px;
            font-weight: 600;
            margin-bottom: 20px;
        }

        /* Info banner */
        .info-banner {
            background-color: #dff3f6;
            padding: 18px;
            border-radius: 10px;
            font-size: 15px;
            margin-bottom: 25px;
        }

        /* Section headers */
        .section-header {
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 10px;
        }

        /* Card container */
        .card {
            background-color: #ffffff;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            margin-bottom: 30px;
        }

        /* Generate button */
        div.stButton > button,
        div.stDownloadButton > button {
            background-color: #aee1ff !important;
            color: #00334d !important;
            font-weight: 600;
            border-radius: 8px;
            height: 46px;
            border: none;
        }

        div.stButton > button:hover {
            background-color: #97d6fb !important;
        }

        /* Sidebar polish only (content unchanged) */
        section[data-testid="stSidebar"] {
            background-color: #f1f3f5;
        }

        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown(
        """
        <div class="info-banner">
        <strong>Welcome to the Playwright Test Runner!</strong><br>
        Paste a function generated by Playwright codegen → get a full runnable test script 
        with step-level pass/fail tracking, Excel report and test case documentation.
        </div>
        """,
        unsafe_allow_html=True
    )

    # ---------------- Sidebar ----------------
    with st.sidebar:
        st.markdown("---")
        st.header("📚 How to Use")
        st.markdown("""
            1. Paste your Playwright codegen function in the text area below.
            2. (Optional) Add any extra context or expected results to help the LLM generate better test cases.
            3. Click 'Generate Runnable Test + Test Cases'.
            4. Download the generated Python script (.py) and test cases (Excel).
            5. Run the script locally to execute the tests and get an execution report in Excel
        """
        )

        st.markdown("---")

        st.markdown("### ℹ️ About")
        st.markdown(
            """
            This tool:

            1. Takes raw codegen output  
            2. Adds detailed step-by-step pass/fail checks  
            3. Generates Excel execution report  
            4. Generates test case documentation in Excel  
            5. Opens browser visibly so you can watch
            """
        )

    # ---------------- Main UI ----------------
    st.markdown('', unsafe_allow_html=True)
    st.markdown('📝 Test Generation', unsafe_allow_html=True)

    with st.form("playwright_test_form"):
        code_input = st.text_area(
            "Paste Playwright Codegen Function",
            placeholder="""def test_example():
    page.goto("https://example.com")
    page.fill("#username", "testuser")
    page.fill("#password", "testpass123")
    page.click("button[type=submit]")
    page.wait_for_selector("text=Dashboard")""",
            height=240,
            help="Paste only the function generated by Playwright codegen (sync API)."
        )

        extra_context = st.text_input(
            "Additional Context (optional)",
            placeholder="URL = https://myapp.com, should see 'Welcome' message after login, check cart count = 2",
            help="Any URL, login info, expected text, assertions, or special notes"
        )

        submitted = st.form_submit_button(
            "🚀 Generate Runnable Test + Test Cases",
            use_container_width=True,
            type="primary"
        )

    st.markdown('', unsafe_allow_html=True)

    # ---------------- Submission Handling ----------------
    if submitted:
        if not code_input.strip():
            st.error("⚠️ Please paste a Playwright codegen function first.")
        else:
            with st.spinner("🔄 Generating instrumented script & test cases..."):

            # -------- 1. Generate runnable script --------
                script_prompt = TRANSFORM_PROMPT.format(
                    input_code=code_input.strip(),
                    extra_context=extra_context.strip()
                    if extra_context else "No extra context provided."
                )
    
                script_agent = GroqAgent(
                    system_prompt="You are a strict Playwright instrumentation engine. Output ONLY valid Python code. No explanations."
                )
    
                script_response = asyncio.run(
                    script_agent.generate(script_prompt)
                )
    
                generated_code = clean_generated_code(script_response)
    
                # -------- 2. Generate test cases --------
                testcase_prompt = TESTCASE_PLAN_PROMPT.format(
                    input_code=code_input.strip(),
                    extra_context=extra_context.strip()
                    if extra_context else "No extra context provided."
                )
    
                testcase_agent = GroqAgent(
                    system_prompt="You generate structured QA test cases based strictly on the provided script."
                )
    
                testcase_response = asyncio.run(
                    testcase_agent.generate(testcase_prompt)
                )


            # -------- Display Generated Script --------
            st.markdown('', unsafe_allow_html=True)
            st.markdown('🎉 Generated Runnable Test Script', unsafe_allow_html=True)

            st.code(generated_code, language="python", line_numbers=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            script_filename = f"playwright_test_{timestamp}.py"

            st.download_button(
                label="📥 Download Test Script (.py)",
                data=generated_code,
                file_name=script_filename,
                mime="text/x-python",
                use_container_width=True
            )

            st.markdown("### 💡 How to run it")
            st.code(
                f"""# Install dependencies (once)
pip install playwright pandas openpyxl
playwright install

# Run the script
python {script_filename}

# Output:
# → Browser opens (visible)
# → Steps logged with PASS/FAIL
# → Excel report: test_results_*.xlsx
""",
                language="bash"
            )

            # -------- Display & Export Test Cases --------
            st.markdown('', unsafe_allow_html=True)
            st.markdown('📋 Generated Test Cases', unsafe_allow_html=True)

            # Uses your existing function
            parse_and_export_testcases(testcase_response)

            # Preview parsed test cases (optional)
            if (
                'test_cases_list' in st.session_state
                and st.session_state.test_cases_list
            ):
                df = pd.DataFrame(st.session_state.test_cases_list)

                st.dataframe(
                    df,
                    column_config={
                        "Step-by-step actions": st.column_config.TextColumn(width="medium"),
                        "Expected Result": st.column_config.TextColumn(width="medium"),
                        "Test Case Description": st.column_config.TextColumn(width="medium"),
                    },
                    use_container_width=True,
                    hide_index=True
                )

                # Download Excel
                try:
                    with open("cleaned_generated_test_cases.xlsx", "rb") as f:
                        st.download_button(
                            label="📥 Download Test Cases Excel",
                            data=f,
                            file_name=f"test_cases_{timestamp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                except FileNotFoundError:
                    st.warning("Excel file was not created successfully.")
            else:
                st.info("No structured test cases detected in response. Raw output:")
                st.code(testcase_response, language="text")

            st.markdown('', unsafe_allow_html=True)
            st.markdown(
                """
                ✅ Done! Download script and run locally. Test cases saved as Excel.
                """,
                unsafe_allow_html=True
            )


if __name__ == "__main__":
    main()






